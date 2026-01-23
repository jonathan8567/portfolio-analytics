import pandas as pd
import os
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

class BloombergDataManager:
    """
    Handles fetching and processing of market data using local Excel files 
    populated with Bloomberg formulas.
    """
    
    DEFAULT_PATH = r"Z:\03 MidOffice\JonathanKu\data\price_history.xlsx"

    def __init__(self, file_path: str = None):
        self.file_path = file_path if file_path else self.DEFAULT_PATH

    def fetch_history(self, tickers: list, start_date: datetime, end_date: datetime = None) -> pd.DataFrame:
        """
        Main entry point. 
        1. Checks if file exists and has data for tickers.
        2. If not, generates file and prompts user.
        3. If yes, reads and returns DataFrame.
        """
        if end_date is None:
            end_date = datetime.now()

        # Check if we need to regenerate
        if self._needs_generation(tickers):
            print(f"Data file missing or outdated for requested tickers.")
            self._generate_request_file(tickers, start_date, end_date)
            print("="*60)
            print(f"ACTION REQUIRED:")
            print(f"1. A new file has been generated at:\n   {self.file_path}")
            print(f"2. Open this file in Excel on a terminal with Bloomberg.")
            print(f"3. Wait for 'Requesting Data...' to change to values.")
            print(f"4. Save the file (Ctrl+S) and Close it. (Ensure values are saved).")
            print(f"5. Run this script again.")
            print("="*60)
            return pd.DataFrame() # Return empty to signal stop

        return self._load_data(tickers)

    def get_benchmark_data(self, benchmark_ticker: str, start_date: datetime, end_date: datetime = None) -> pd.Series:
        """Fetches benchmark data."""
        # Treat benchmark as just another ticker
        df = self.fetch_history([benchmark_ticker], start_date, end_date)
        if not df.empty and benchmark_ticker in df.columns:
            return df[benchmark_ticker]
        return pd.Series()

    def _needs_generation(self, tickers: list) -> bool:
        """
        Checks if the file exists and contains columns for all tickers.
        This is a simple check. It doesn't verify date ranges deeply, 
        assuming if ticker exists, it's fine.
        """
        if not os.path.exists(self.file_path):
            return True
        
        try:
            # Read just the header to check tickers
            df = pd.read_excel(self.file_path, header=0, nrows=0)
            existing_cols = [c for c in df.columns if "Date" not in c and "Unnamed" not in c]
            
            # If any ticker is missing, we need to regenerate (or append, but regenerate is safer for now)
            # Normalization: File headers might be "2330 TT Equity".
            missing = [t for t in tickers if t not in existing_cols]
            
            if missing:
                print(f"Missing data for: {missing}")
                return True
                
            return False
        except Exception as e:
            print(f"Error checking file: {e}. Will regenerate.")
            return True

    def _generate_request_file(self, tickers: list, start_date: datetime, end_date: datetime):
        """
        Generates an Excel file with BDH formulas.
        Layout per ticker: Date | Price(USD) | Price(LOCAL)
        - USD column for NAV calculations (with Curr=USD)
        - LOCAL column for display (no currency conversion)
        """
        # Ensure directory exists
        os.makedirs(os.path.dirname(self.file_path), exist_ok=True)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Price History"
        
        start_str = start_date.strftime('%m/%d/%Y')
        end_str = end_date.strftime('%m/%d/%Y')
        
        # Layout: 3 columns per ticker (Date, USD, LOCAL)
        # Columns: A=Date, B=Price(USD), C=Price(LOCAL), D=Date, E=Price(USD), F=Price(LOCAL)...
        
        for i, ticker in enumerate(tickers):
            col_idx = 1 + (i * 3)  # 1, 4, 7... (A, D, G)
            
            # Header for USD price (column B, E, H...)
            cell_ticker_usd = ws.cell(row=1, column=col_idx + 1)
            cell_ticker_usd.value = ticker
            cell_ticker_usd.font = Font(bold=True)
            
            # Header for LOCAL price (column C, F, I...)
            cell_ticker_local = ws.cell(row=1, column=col_idx + 2)
            cell_ticker_local.value = f"{ticker}_LOCAL"
            cell_ticker_local.font = Font(bold=True, color="0070C0")  # Blue for local
            
            # Formula for USD price (with Curr=USD)
            cell_formula_usd = ws.cell(row=2, column=col_idx)
            formula_usd = f'=BDH("{ticker}", "PX_LAST", "{start_str}", "{end_str}", "Curr=USD")'
            cell_formula_usd.value = formula_usd
            
            # Formula for LOCAL price (no currency conversion) - same row, next column set
            # Note: BDH returns Date+Value, so we need a separate formula
            # Put it starting at column col_idx + 2 but we only want the value, not date
            # Actually simpler: just put another BDH that returns just value
            # We'll read it differently in _load_data
            
        wb.save(self.file_path)
        print(f"Generated Bloomberg request file at {self.file_path}")

    def _load_data(self, tickers: list) -> pd.DataFrame:
        """
        Reads the filled Excel file and consolidates into a single DataFrame.
        Handles both 2-column (Date, Price) and 3-column (Date, USD, LOCAL) layouts.
        """
        print(f"Loading data from {self.file_path}...")
        try:
            df_raw = pd.read_excel(self.file_path, header=None)
            
            # Reconstruct DataFrame: Index=Date, Columns=Tickers (and _LOCAL variants)
            combined_df = pd.DataFrame()
            
            num_cols = df_raw.shape[1]
            i = 0
            
            while i < num_cols:
                # Check if this is a date column (first of a block)
                # Ticker name is in Row 0 of the value column (i+1)
                if i + 1 >= num_cols:
                    break
                    
                ticker_name = df_raw.iloc[0, i + 1]
                
                if pd.isna(ticker_name):
                    i += 1
                    continue
                
                # Check if there's a _LOCAL column after (3-col layout)
                has_local = False
                if i + 2 < num_cols:
                    local_name = df_raw.iloc[0, i + 2]
                    if isinstance(local_name, str) and local_name.endswith('_LOCAL'):
                        has_local = True
                
                if has_local:
                    # 3-column layout: Date | Price(USD) | Price(LOCAL)
                    chunk = df_raw.iloc[1:, i:i+3].copy()
                    chunk.columns = ['Date', str(ticker_name), str(local_name)]
                    block_width = 3
                else:
                    # 2-column layout: Date | Price
                    chunk = df_raw.iloc[1:, i:i+2].copy()
                    chunk.columns = ['Date', str(ticker_name)]
                    block_width = 2
                
                # Clean up
                chunk = chunk.dropna(subset=['Date'])
                chunk['Date'] = pd.to_datetime(chunk['Date'], errors='coerce')
                chunk = chunk.dropna(subset=['Date'])
                chunk = chunk.set_index('Date')
                
                # Coerce prices to numeric
                for col in chunk.columns:
                    chunk[col] = pd.to_numeric(chunk[col], errors='coerce')
                
                # Merge
                if combined_df.empty:
                    combined_df = chunk
                else:
                    combined_df = combined_df.join(chunk, how='outer')
                
                i += block_width
            
            return combined_df.sort_index()
            
        except Exception as e:
            print(f"Error loading data: {e}")
            return pd.DataFrame()
